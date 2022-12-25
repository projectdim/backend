from typing import List, Dict, Generator

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.utils.geocoding import geocode_address
from app.db.session import SessionLocal
from app.models.location import Location
from app.crud.crud_geospatial import create_index
from app.crud.crud_user import get_by_email
from app.schemas.location import LocationReports
from app.core.config import settings
from app.crud.crud_location import get_location_by_coordinates, submit_location_reports


def serialize_spreadsheet(spreadsheet, sheet_type: int) -> List[Dict]:
    locations = []
    if sheet_type == 1:
        for row in spreadsheet.iter_rows(min_row=2, min_col=2):
            if row[0].value is None:
                continue
            location = {
                "address": row[0].value,
                "city": row[1].value,
                "country": row[2].value,
                "index": row[3].value,
                "reports": {
                    "buildingCondition": {
                        "flag": row[4].value.lower(),
                        "description": row[10].value if row[10].value else ""
                    },
                    "electricity": {
                        "flag": row[5].value.lower(),
                        "description": ""
                    },
                    "carEntrance": {
                        "flag": row[6].value.lower(),
                        "description": ""
                    },
                    "water": {
                        "flag": row[7].value.lower(),
                        "description": ""
                    },
                    "fuelStation": {
                        "flag": row[8].value.lower(),
                        "description": ""
                    },
                    "hospital": {
                        "flag": row[9].value.lower(),
                        "description": ""
                    }
                }
            }
            if location["address"] is None:
                continue
            locations.append(location)

    elif sheet_type == 2:
        for row in spreadsheet.iter_rows(min_row=2, min_col=1):
            if row[0].value is None:
                continue
            location = {
                "address": row[0].value,
                "street_number": int(row[1].value) if isinstance(row[1].value, float) else row[1].value,
                "city": row[2].value,
                "country": row[3].value,
                "index": row[4].value,
                "reports": {
                    "buildingCondition": {
                        "flag": row[5].value.lower(),
                        "description": row[11].value if row[11].value else ""
                    },
                    "electricity": {
                        "flag": row[6].value.lower(),
                        "description": ""
                    },
                    "carEntrance": {
                        "flag": row[7].value.lower(),
                        "description": ""
                    },
                    "water": {
                        "flag": row[8].value.lower(),
                        "description": ""
                    },
                    "fuelStation": {
                        "flag": row[9].value.lower(),
                        "description": ""
                    },
                    "hospital": {
                        "flag": row[10].value.lower(),
                        "description": ""
                    }
                }
            }
            if location["address"] is None:
                continue
            locations.append(location)
    return locations


def geocode_locations(locations: List[Dict]):

    geocoded_locations = []
    for location in locations:
        addr_str = '{}'.format(location["address"] + " " + str(location.get('street_number')) if location.get('street_number',None) else location["address"])
        coordinates = geocode_address(addr_str, location["city"])
        if not coordinates:
            continue
        location["lat"] = coordinates.latitude
        location["lng"] = coordinates.longitude
        geocoded_locations.append(location)

    return geocoded_locations


def add_to_db(location):
    try:
        db: Session = SessionLocal()
        existing_location = get_location_by_coordinates(db, lat=location.get('lat'), lng=location.get('lng'))
        if existing_location:
            return None
        reporting_user = get_by_email(db, email=settings.FIRST_SUPERUSER)
        db_obj = Location(
            address=location.get('address'),
            index=location.get('index'),
            lat=location.get('lat'),
            lng=location.get('lng'),
            country=location.get('country'),
            city=location.get('city'),
            status=3,
            reports=location.get('reports'),
            street_number=location.get('street_number', None)
        )

        db.add(db_obj)
        db.commit()
        db.refresh(db_obj)

        index = create_index(db, location_id=db_obj.id, lat=db_obj.lat, lng=db_obj.lng, status=db_obj.status)

        submit_location_reports(
            db,
            obj_in=LocationReports(location_id=db_obj.id, **location.get('reports')),
            user_id=reporting_user.id
        )

        db.close()

        return db_obj

    except Exception as e:
        print(e)
        return None


def bulk_create(spreadsheet_path: str, sheet_type: int):
    try:

        wb2 = load_workbook(spreadsheet_path)
        sheet = wb2.active
        locations = serialize_spreadsheet(sheet, sheet_type=sheet_type)
        print('Finished serializing, total locations {}'.format(len(locations)))
        updated_locations = geocode_locations(locations)
        print('Finished geocoding, total locations {}'.format(len(updated_locations)))
        unregistered_locations = []
        for location in updated_locations:
            if not location.get('lat') or not location.get('lng'):
                continue
            loc = add_to_db(location)
            if not loc:
                unregistered_locations.append(location)

        print('Could not register {} locations.'.format(len(unregistered_locations)))

        return unregistered_locations

    except Exception as e:
        print(e)
        return None
